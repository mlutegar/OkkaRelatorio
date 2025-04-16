import openpyxl
from django.db.models import Sum
from django.http import HttpResponse
from django.views import View
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from rest_framework import viewsets, generics, status
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView

from backend.serializers import UserSerializer, RelatorioSerializer, CadastroSerializer

from backend.models import User, Relatorio


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [AllowAny]
    lookup_field = 'username'

    def perform_create(self, serializer):
        serializer.save(creator=self.request.user)


class RelatorioPorUsuarioView(generics.ListAPIView):
    serializer_class = RelatorioSerializer
    permission_classes = [AllowAny]  # Defina a permissão que você preferir

    def get_queryset(self):
        username = self.kwargs['username']  # Pega o username dos parâmetros de URL
        return Relatorio.objects.filter(colaborador__username=username)  # Filtra os relatórios pelo colaborador

class RelatorioPorSetorView(generics.ListAPIView):
    serializer_class = RelatorioSerializer
    permission_classes = [AllowAny]  # Defina a permissão que você preferir

    def get_queryset(self):
        setor = self.kwargs['setor']  # Pega o setor dos parâmetros de URL
        return Relatorio.objects.filter(setor=setor)  # Filtra os relatórios pelo setor

class CadastroView(generics.CreateAPIView):
    serializer_class = CadastroSerializer
    permission_classes = [AllowAny]  # Permite que qualquer pessoa se cadastre


class RelatorioViewSet(viewsets.ModelViewSet):
    queryset = Relatorio.objects.all()
    serializer_class = RelatorioSerializer
    permission_classes = [AllowAny]

    # Explicitly define all HTTP methods
    def get_method_actions(self, request):
        allowed_methods = {
            'GET': self.list,
            'POST': self.create,
            'PUT': self.update,
            'PATCH': self.partial_update,
            'DELETE': self.destroy
        }
        return allowed_methods

    @action(detail=True, methods=['put', 'patch'], url_path='custom-update')
    def custom_update(self, request, pk=None):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)

        try:
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(serializer.data)
        except Exception as e:
            return Response({
                'error': str(e),
                'request_method': request.method,
                'request_data': request.data
            }, status=status.HTTP_400_BAD_REQUEST)


def export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatórios por Colaborador"

    relatorios_por_colaborador = Relatorio.objects.values('colaborador__first_name').annotate(
        total_horas=Sum('hora_modificada')
    )

    # Cabeçalhos
    headers = ['Nome do Colaborador', 'Total de Horas Modificadas']
    for col_num, header in enumerate(headers, 1):
        ws[f'{get_column_letter(col_num)}1'] = header
        ws[f'{get_column_letter(col_num)}1'].font = openpyxl.styles.Font(bold=True)

    # Pega dados do modelo
    for idx, item in enumerate(relatorios_por_colaborador, start=2):
        ws[f'A{idx}'] = item['colaborador__first_name']
        ws[f'B{idx}'] = item['total_horas']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="relatorio.xlsx"'

    wb.save(response)
    return response